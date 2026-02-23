"""
pipeline.py

This file is a thin wrapper around YOUR working OpenAlex/GitHub pipeline.

✅ Goal: Keep your pipeline logic + comments + POS/NEG terms EXACTLY the same.
✅ We only:
  - accept parameters from the local backend
  - set globals your script already uses (SEED_WORK_IDS, etc.)
  - call your main()
  - return the XLSX output path

INSTRUCTIONS:
1) Paste your entire working script BELOW the "PASTE YOUR SCRIPT HERE" marker.
2) Do NOT change your POS/NEG/ENG_POS/ENG_NEG lists or comments.
3) Ensure your script exports an XLSX file named OUTPUT_XLSX (set below).
"""

import os
from typing import Tuple, List, Callable, Dict, Any, Optional

# Your script should export this exact XLSX filename:
OUTPUT_XLSX = "openalex_ml_perf_candidates_multiseed.xlsx"

# NEW: progress callback storage (set by backend per run)
PROGRESS_CB: Optional[Callable[[Dict[str, Any]], None]] = None


def run_pipeline_to_xlsx(
    seed_work_ids: List[str],
    max_citing_works_per_seed: int,
    author_work_year_range: Tuple[int, int],
    max_works_per_author: int,
    sleep_seconds: float,
    include_seed_authors: bool,
    seed_author_bonus: int,
    eng_gate_enabled: bool,
    eng_gate_min_score: int,
    eng_gate_max_works_scanned: int,
    github_sniff_enabled: bool,
    progress_cb: Optional[Callable[[Dict[str, Any]], None]] = None,
) -> str:
    """
    Called by FastAPI.
    We set the same globals your script already uses, then call main().
    Returns absolute path to XLSX output.
    """

    # NEW: store callback globally so your script can call it
    global PROGRESS_CB
    PROGRESS_CB = progress_cb

    # --- Set your script globals (these names MUST match your script) ---
    global SEED_WORK_IDS
    global MAX_CITING_WORKS_PER_SEED
    global AUTHOR_WORK_YEAR_RANGE
    global MAX_WORKS_PER_AUTHOR
    global SLEEP_SECONDS

    global INCLUDE_SEED_AUTHORS
    global SEED_AUTHOR_BONUS

    global ENG_GATE_ENABLED
    global ENG_GATE_MIN_SCORE
    global ENG_GATE_MAX_WORKS_SCANNED

    # Optional toggle if you have GitHub sniff in-script:
    global GITHUB_SNIFF_ENABLED

    # Mirror your existing tunable variable names
    SEED_WORK_IDS = seed_work_ids
    MAX_CITING_WORKS_PER_SEED = max_citing_works_per_seed
    AUTHOR_WORK_YEAR_RANGE = author_work_year_range
    MAX_WORKS_PER_AUTHOR = max_works_per_author
    SLEEP_SECONDS = sleep_seconds

    INCLUDE_SEED_AUTHORS = include_seed_authors
    SEED_AUTHOR_BONUS = seed_author_bonus

    ENG_GATE_ENABLED = eng_gate_enabled
    ENG_GATE_MIN_SCORE = eng_gate_min_score
    ENG_GATE_MAX_WORKS_SCANNED = eng_gate_max_works_scanned

    # If your script uses this, it will honor it. If not, no harm.
    GITHUB_SNIFF_ENABLED = github_sniff_enabled

    # --- Run your pipeline ---
    main()

    # --- Return output path ---
    out_path = os.path.abspath(OUTPUT_XLSX)
    if not os.path.exists(out_path):
        raise FileNotFoundError(
            f"Expected XLSX output not found: {out_path}. "
            f"Make sure your script exports to {OUTPUT_XLSX}."
        )
    return out_path


# ======================================================================
# ======================= PASTE YOUR SCRIPT HERE ========================
# ======================================================================
import requests
import time
import csv
from collections import defaultdict, Counter

# NEW: Excel export + hyperlinks
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from urllib.parse import quote_plus

# NEW: GitHub sniff (optional token for higher rate limits)
import os
from getpass import getpass
import re
from datetime import datetime, timezone


MAILTO = os.getenv("OPENALEX_EMAIL", "your.email@example.com")
BASE = "https://api.openalex.org"


GITHUB_TOKEN = os.getenv("GITHUB_TOKEN")

# ----------------------------
# TUNABLES
# ----------------------------
SEED_WORK_IDS = [
    "W4318541647",   # replace / add more seeds
   # "W2804032941",
    # "W3153553004",
]
MAX_CITING_WORKS_PER_SEED = 300   # per seed cap
AUTHOR_WORK_YEAR_RANGE = (2019, 2026)
MAX_WORKS_PER_AUTHOR = 400
SLEEP_SECONDS = 0.10

# NEW: Tier-0 (seed authors) tunables
INCLUDE_SEED_AUTHORS = True
SEED_AUTHOR_BONUS = 3  # how much to upweight seed authors in author_counter

# Engineering-gate tunables
ENG_GATE_ENABLED = True
ENG_GATE_MIN_SCORE = 0
ENG_GATE_MAX_WORKS_SCANNED = 30

# ----------------------------
# KEYWORDS (role scoring)
# ----------------------------
POS = [
    # Research → Production ML infrastructure
    "training infrastructure", "inference infrastructure", "production machine learning",
    "ml platform", "end-to-end pipeline", "model serving", "deployment",
    "scalable training", "distributed training", "experiment management", "mlops",

    # Mathematical & optimization depth
    "optimization", "optimization theory", "regularization", "linear algebra",
    "convex optimization", "non-convex optimization", "stochastic optimization",
    "gradient methods", "numerical optimization", "theoretical analysis",

    # ML frameworks & internals (kept; bucket below broadened)
    "pytorch", "jax", "tensorflow", "autograd", "automatic differentiation",
    "xla", "jit", "jit compilation", "custom operators",

    # Research code quality & reproducibility
    "reproducible research", "research codebase", "experimental framework",
    "experiment tracking", "ablation study", "benchmarking",
    "software engineering", "testing", "modular code",

    # State-of-the-art engagement & inventive thinking
    "state of the art", "sota", "novel approach", "empirical study",
    "comparative analysis", "failure modes", "trade-offs",
    "design decisions", "alternative approach"
]

NEG = [
    # ----------------------------
    # Eval-only / leaderboard-heavy / surveys
    # ----------------------------
    "benchmark", "benchmarks", "leaderboard", "leaderboards",
    "mmlu", "glue", "superglue", "imagenet", "coco",
    "human evaluation", "user evaluation", "crowdsourced evaluation",
    "survey paper", "survey", "systematic review", "meta-analysis",
    "empirical comparison only",

    # ----------------------------
    # Prompting / alignment / UX-level LLM work
    # ----------------------------
    "prompt engineering", "prompt tuning", "instruction tuning",
    "in-context learning", "few-shot prompting",
    "rlhf", "alignment", "value alignment",
    "chatgpt", "gpt-", "llm alignment",
    "preference modeling", "human feedback",

    # ----------------------------
    # Pure theory / math-heavy (non-systems)
    # ----------------------------
    "theorem", "proof", "lemma", "corollary",
    "theoretical analysis only", "theoretical guarantee",
    "convergence proof", "asymptotic analysis",
    "closed-form solution",

    # ----------------------------
    # Cryptography / privacy (major source of drift)
    # ----------------------------
    "cryptography", "cryptographic", "encryption", "encrypted",
    "homomorphic", "fully homomorphic", "fhe",
    "secure computation", "multi-party computation", "mpc",
    "zero-knowledge", "zkp", "zk-snark",
    "privacy-preserving", "confidential computing",
    "trusted execution", "secure enclave", "tee",

    # ----------------------------
    # Narrow domain application ML (non-transferable)
    # ----------------------------
    "medical imaging", "radiology", "clinical study",
    "bioinformatics", "genomics", "proteomics",
    "healthcare application", "drug discovery",
    "remote sensing", "earth observation",
    "smart grid", "traffic prediction",
    "agricultural", "crop yield",

    # ----------------------------
    # Social science / HCI / education
    # ----------------------------
    "user study", "usability study", "human factors",
    "hci", "education", "learning sciences",
    "survey data", "questionnaire", "interviews",

    # ----------------------------
    # Policy / governance / standards
    # ----------------------------
    "ethics", "fairness", "bias mitigation",
    "ai governance", "policy framework",
    "iso", "iec", "standards committee",
    "regulatory compliance",

    # ----------------------------
    # Non-engineering signals
    # ----------------------------
    "opinion piece", "position paper",
    "vision paper", "perspective",
]

# ----------------------------
# Engineering-gate keywords
# ----------------------------
ENG_POS = [
    "system", "systems", "infrastructure", "platform", "pipeline", "serving", "deployment",
    "distributed", "scalable", "throughput", "latency", "performance", "optimization",
    "compiler", "runtime", "api", "framework", "mlops", "monitoring",
    "reproducible", "benchmark", "engineering", "software", "toolkit", "library",
    "open source", "github"
]

ENG_NEG = [
    "theorem", "proof", "lemma", "corollary", "lower bound", "upper bound",
    "asymptotic", "we prove", "convergence proof", "formal verification",
    "theoretical", "theoretical guarantee"
]


# ----------------------------
# NEW: Progress reporting (real % in UI)
# ----------------------------
def report_progress(stage: str, processed: int, total: int, message: str = ""):
    """
    Sends progress to the FastAPI backend via the PROGRESS_CB set in run_pipeline_to_xlsx().
    overall_pct is computed across major stages:
      0: citing works
      1: authors
      2: github sniff (optional)
      3: export
    """
    try:
        cb = globals().get("PROGRESS_CB")
        if not cb:
            return

        total = max(int(total), 1)
        processed = max(int(processed), 0)
        processed = min(processed, total)

        stage_pct = (processed / total) * 100.0

        stage_order = ["citing_works", "authors", "github_sniff", "export"]
        if stage not in stage_order:
            stage_index = 0
            stages_count = len(stage_order)
        else:
            stage_index = stage_order.index(stage)
            stages_count = len(stage_order)

        overall_pct = ((stage_index + (processed / total)) / stages_count) * 100.0

        cb({
            "stage": stage,
            "processed": processed,
            "total": total,
            "stage_pct": round(stage_pct, 2),
            "overall_pct": round(overall_pct, 2),
            "message": message or stage,
        })
    except Exception:
        return


def get(url, params=None):
    params = params or {}
    params["mailto"] = MAILTO
    r = requests.get(url, params=params, timeout=30)
    r.raise_for_status()
    time.sleep(SLEEP_SECONDS)
    return r.json()


def score_text(text: str):
    t = (text or "").lower()
    pos = sum(1 for k in POS if k in t)
    neg = sum(1 for k in NEG if k in t)
    return pos - neg, pos, neg


def text_for_work(work):
    title = (work.get("display_name") or "").lower()
    concepts = " ".join([(c.get("display_name", "") or "") for c in work.get("concepts", [])]).lower()
    venue = (work.get("primary_location", {}) or {}).get("source", {}) or {}
    venue_name = (venue.get("display_name") or "").lower()
    return f"{title} {concepts} {venue_name}"


def extract_author_ids_from_work(work):
    author_ids = []
    for a in work.get("authorships", []):
        author = a.get("author") or {}
        if author.get("id"):
            author_ids.append(author["id"].split("/")[-1])  # Ax...
    # de-dupe, keep order
    seen = set()
    out = []
    for x in author_ids:
        if x not in seen:
            seen.add(x)
            out.append(x)
    return out


def fetch_author(author_id):
    return get(f"{BASE}/authors/{author_id}")


def fetch_author_works(author_id, year_start, year_end, max_works):
    works = []
    cursor = "*"
    while True:
        data = get(
            f"{BASE}/works",
            params={
                "filter": f"authorships.author.id:{author_id},publication_year:{year_start}-{year_end}",
                "per_page": 200,
                "cursor": cursor,
                "sort": "cited_by_count:desc",
                "select": "id,display_name,publication_year,cited_by_count,concepts,primary_location",
            },
        )
        works.extend(data.get("results", []))
        cursor = data.get("meta", {}).get("next_cursor")
        if not cursor or len(works) >= max_works:
            break
    return works[:max_works]


def fetch_citing_works(seed_work_id, max_works):
    """Returns a list of Work IDs (W...) for papers that cite the seed."""
    citing_ids = []
    cursor = "*"
    while True:
        data = get(
            f"{BASE}/works",
            params={
                "filter": f"cites:{seed_work_id}",
                "per_page": 200,
                "cursor": cursor,
                "select": "id,display_name,publication_year,cited_by_count",
            },
        )
        for w in data.get("results", []):
            wid = (w.get("id") or "").split("/")[-1]
            if wid.startswith("W"):
                citing_ids.append(wid)

        cursor = data.get("meta", {}).get("next_cursor")
        if not cursor or len(citing_ids) >= max_works:
            break

    citing_ids = list(dict.fromkeys(citing_ids))
    return citing_ids[:max_works]


def fetch_work_full(work_id):
    return get(f"{BASE}/works/{work_id}")


def make_linkedin_xray(name, org=None):
    bias = '(CUDA OR GPU OR "deep learning compiler" OR runtime OR "kernel fusion" OR "distributed training" OR NCCL OR Triton OR ONNX OR TensorRT OR TVM OR XLA OR MLIR)'
    if org:
        return f'site:linkedin.com/in "{name}" ("{org}" OR {bias})'
    return f'site:linkedin.com/in "{name}" {bias}'


def eng_score_for_author(works):
    sample = works[:ENG_GATE_MAX_WORKS_SCANNED]
    text_blob = " ".join(text_for_work(w) for w in sample).lower()
    pos = sum(1 for k in ENG_POS if k in text_blob)
    neg = sum(1 for k in ENG_NEG if k in text_blob)
    return pos - neg


# NEW: Google search URL for name (used for Excel hyperlink)
def make_google_name_search(name: str) -> str:
    q = quote_plus(f"\"{name}\"")
    return f"https://www.google.com/search?q={q}"


# ----------------------------
# NEW: GitHub sniff test (runs only for authors who pass EngGate)
# ----------------------------
GITHUB_SNIFF_ENABLED = True  # set by wrapper
GITHUB_TOKEN = os.environ.get("GITHUB_TOKEN", "").strip()
GITHUB_BASE = "https://api.github.com"
GITHUB_SLEEP_SECONDS = 0.15  # keep polite; GitHub also rate-limits

INDUSTRY_POS_WORDS = [
    "engineer", "software", "swe", "developer", "ml engineer", "research engineer",
    "platform", "infrastructure", "systems", "backend", "compiler", "runtime",
    "nvidia", "google", "meta", "facebook", "amazon", "aws", "microsoft", "openai",
    "apple", "netflix", "uber", "lyft", "databricks", "deepmind", "anthropic",
    "intel", "amd", "tesla", "stripe", "cloud", "production"
]

ACADEMIA_NEG_WORDS = [
    "professor", "assistant professor", "associate professor", "full professor",
    "phd student", "ph.d", "postdoc", "post-doctoral", "postdoctoral",
    "faculty", "department", "university", "institute", "laboratory", "lab"
]

# Known “industry-looking” GitHub orgs (bonus signal if they contribute to these repos/orgs)
INDUSTRY_GH_ORGS = [
    "pytorch", "triton-lang", "tensorflow", "jax-ml", "openai", "nvidia",
    "google", "microsoft", "meta", "facebookresearch", "amazon-science",
    "apache", "llvm", "onnx", "mlir", "tvm", "huggingface"
]

def gh_headers():
    h = {
        "Accept": "application/vnd.github+json",
        "User-Agent": "openalex-recruiting-script",
    }
    if GITHUB_TOKEN:
        h["Authorization"] = f"Bearer {GITHUB_TOKEN}"
    return h

def gh_get(url, params=None):
    r = requests.get(url, headers=gh_headers(), params=params or {}, timeout=30)
    if r.status_code == 403:
        return {"__error__": "forbidden", "__status__": 403, "__text__": r.text}
    if r.status_code == 404:
        return {"__error__": "not_found", "__status__": 404, "__text__": r.text}
    r.raise_for_status()
    time.sleep(GITHUB_SLEEP_SECONDS)
    return r.json()

def normalize_name_for_search(name: str) -> str:
    return re.sub(r"\s+", " ", (name or "").strip())

def parse_github_dt(dt_str: str):
    if not dt_str:
        return None
    try:
        return datetime.strptime(dt_str, "%Y-%m-%dT%H:%M:%SZ").replace(tzinfo=timezone.utc)
    except Exception:
        return None

def gh_industry_signal_from_user(user_json: dict) -> (int, str):
    """
    Score a GitHub user profile as "industry-leaning" based on bio/company + repo stats.
    Returns (signal_score, notes_string).
    """
    score = 0
    notes = []

    login = user_json.get("login", "")
    html_url = user_json.get("html_url", "")
    bio = (user_json.get("bio") or "").lower()
    company = (user_json.get("company") or "").lower()
    followers = int(user_json.get("followers") or 0)
    public_repos = int(user_json.get("public_repos") or 0)
    updated_at = parse_github_dt(user_json.get("updated_at"))

    blob = f"{bio} {company}".strip()

    pos_hits = [w for w in INDUSTRY_POS_WORDS if w in blob]
    neg_hits = [w for w in ACADEMIA_NEG_WORDS if w in blob]
    if pos_hits:
        score += min(len(pos_hits), 6)
        notes.append(f"bio/company pos: {', '.join(pos_hits[:6])}")
    if neg_hits:
        score -= min(len(neg_hits), 6)
        notes.append(f"bio/company acad: {', '.join(neg_hits[:6])}")

    if followers >= 100:
        score += 2
        notes.append("followers>=100")
    elif followers >= 30:
        score += 1
        notes.append("followers>=30")

    if public_repos >= 30:
        score += 2
        notes.append("public_repos>=30")
    elif public_repos >= 10:
        score += 1
        notes.append("public_repos>=10")

    if updated_at:
        days = (datetime.now(timezone.utc) - updated_at).days
        if days <= 365:
            score += 2
            notes.append("active<=12mo")
        elif days <= 730:
            score += 1
            notes.append("active<=24mo")

    repos = gh_get(f"{GITHUB_BASE}/users/{login}/repos", params={"per_page": 50, "sort": "updated"})
    if isinstance(repos, dict) and repos.get("__error__"):
        notes.append("repos_fetch_failed")
    else:
        max_stars = 0
        total_stars_top = 0
        recent_push = None
        org_bonus = 0

        for repo in repos[:50]:
            st = int(repo.get("stargazers_count") or 0)
            max_stars = max(max_stars, st)
            total_stars_top += st

            pushed = parse_github_dt(repo.get("pushed_at"))
            if pushed and (recent_push is None or pushed > recent_push):
                recent_push = pushed

            full_name = (repo.get("full_name") or "").lower()
            for org in INDUSTRY_GH_ORGS:
                if full_name.startswith(org.lower() + "/"):
                    org_bonus = 1
                    break

        if max_stars >= 200:
            score += 3
            notes.append("max_repo_stars>=200")
        elif max_stars >= 50:
            score += 2
            notes.append("max_repo_stars>=50")
        elif max_stars >= 20:
            score += 1
            notes.append("max_repo_stars>=20")

        if total_stars_top >= 300:
            score += 2
            notes.append("top_repos_total_stars>=300")
        elif total_stars_top >= 100:
            score += 1
            notes.append("top_repos_total_stars>=100")

        if recent_push:
            days = (datetime.now(timezone.utc) - recent_push).days
            if days <= 180:
                score += 2
                notes.append("recent_push<=6mo")
            elif days <= 365:
                score += 1
                notes.append("recent_push<=12mo")

        if org_bonus:
            score += 1
            notes.append("known_org_repo")

    if html_url:
        notes.append(f"profile:{html_url}")

    return score, " | ".join(notes[:8])

def github_sniff_for_author(name: str):
    """
    Returns:
      github_found (bool)
      github_industry_signal (int)
      github_notes (str)
    """
    nm = normalize_name_for_search(name)
    if not nm:
        return False, 0, "no_name"

    q = f'"{nm}" in:fullname'
    data = gh_get(f"{GITHUB_BASE}/search/users", params={"q": q, "per_page": 5})

    if isinstance(data, dict) and data.get("__error__"):
        if data.get("__status__") == 403:
            return False, 0, "github_rate_limited_or_forbidden"
        return False, 0, f"github_search_failed:{data.get('__status__')}"

    items = data.get("items", []) if isinstance(data, dict) else []
    if not items:
        return False, 0, "no_github_match"

    best_score = None
    best_notes = ""
    best_login = None

    for item in items[:5]:
        login = item.get("login")
        if not login:
            continue

        user = gh_get(f"{GITHUB_BASE}/users/{login}")
        if isinstance(user, dict) and user.get("__error__"):
            continue

        score, notes = gh_industry_signal_from_user(user)
        followers = int(user.get("followers") or 0)

        if best_score is None or score > best_score or (score == best_score and followers > 0):
            best_score = score
            best_notes = f"login:{login} | {notes}"
            best_login = login

    if best_score is None:
        return False, 0, "github_user_fetch_failed"

    return True, int(best_score), best_notes


def main():
    # 0) NEW: Tier-0 candidate collection (seed authors)
    author_counter = Counter()
    author_to_example_works = defaultdict(list)

    if INCLUDE_SEED_AUTHORS:
        for seed in SEED_WORK_IDS:
            try:
                seed_work = fetch_work_full(seed)
            except Exception as e:
                print(f"WARNING: failed to fetch seed work {seed}: {e}")
                continue

            seed_title = seed_work.get("display_name", "")
            seed_author_ids = extract_author_ids_from_work(seed_work)

            for aid in seed_author_ids:
                author_counter[aid] += SEED_AUTHOR_BONUS  # bonus so seed authors rank
                if len(author_to_example_works[aid]) < 3 and seed_title:
                    author_to_example_works[aid].append(f"[SEED] {seed_title}")

        print(f"Tier-0: added seed authors from {len(SEED_WORK_IDS)} seeds (bonus={SEED_AUTHOR_BONUS})")

    # 1) Expand seeds -> union of citing works
    citing_work_ids = set()
    seed_to_citing_count = {}

    for seed in SEED_WORK_IDS:
        ids = fetch_citing_works(seed, MAX_CITING_WORKS_PER_SEED)
        seed_to_citing_count[seed] = len(ids)
        citing_work_ids.update(ids)
        print(f"Seed {seed}: {len(ids)} citing works")

    citing_work_ids = list(citing_work_ids)
    print(f"Total unique citing works across seeds: {len(citing_work_ids)}")

    report_progress("citing_works", 0, len(citing_work_ids), "Collecting citing works…")

    # 2) Extract authors from citing works
    for i, wid in enumerate(citing_work_ids, 1):
        work = fetch_work_full(wid)
        author_ids = extract_author_ids_from_work(work)

        for aid in author_ids:
            author_counter[aid] += 1
            if len(author_to_example_works[aid]) < 3:
                title = work.get("display_name", "")
                if title:
                    author_to_example_works[aid].append(title)

        if i % 25 == 0:
            print(f"Processed {i}/{len(citing_work_ids)} citing works...")
            report_progress("citing_works", i, len(citing_work_ids), f"Processed citing works: {i}/{len(citing_work_ids)}")

    # Keep only authors that appear in >=2 citing papers OR are tier-0 seed authors
    # (Tier-0 authors have bonus counts; still, to be explicit, we keep anyone >=2 OR in counter via seed bonus)
    author_pool = [aid for aid, cnt in author_counter.items() if cnt >= 2]
    print(f"Author pool size (cnt >= 2, incl tier-0 bonuses): {len(author_pool)}")
    if not author_pool:
        author_pool = list(author_counter.keys())
        print(f"Falling back to all authors from graph (tier-0 + citing authors): {len(author_pool)}")

    report_progress("authors", 0, len(author_pool), "Scoring authors…")

    # 3) Score each author by their own recent works
    rows = []
    year_start, year_end = AUTHOR_WORK_YEAR_RANGE

    kept = 0
    skipped_by_gate = 0
    skipped_by_artifacts = 0

    # NEW: GitHub sniff counters (visibility)
    github_checked = 0
    github_found_count = 0
    github_total_est = len(author_pool)

    for idx, aid in enumerate(sorted(author_pool), 1):
        author = fetch_author(aid)
        name = author.get("display_name", "")

        insts = author.get("last_known_institutions") or []
        org = insts[0].get("display_name") if insts else ""

        works = fetch_author_works(aid, year_start, year_end, MAX_WORKS_PER_AUTHOR)
        if not works:
            continue

        # Engineering gate
        if ENG_GATE_ENABLED:
            eng_score = eng_score_for_author(works)
            if eng_score < ENG_GATE_MIN_SCORE:
                skipped_by_gate += 1
                continue
        else:
            eng_score = None

        bucket_hits = {
            "infra_prod": 0,
            "math_optimization": 0,
            "ml_frameworks": 0,
            "research_workflows": 0,
            "inventive_thinking": 0,
            "software_artifacts": 0,
        }
        scored_works = []

        for w in works:
            t = text_for_work(w)
            s, pos, neg = score_text(t)
            scored_works.append((
                s,
                w.get("cited_by_count", 0),
                w.get("publication_year", None),
                w.get("display_name", ""),
                w.get("id", ""),
            ))

            tc = t  # already lower-cased text

            if any(k in tc for k in [
                "training infrastructure", "inference infrastructure", "ml platform",
                "model serving", "deployment", "production machine learning",
                "end-to-end pipeline", "mlops", "scalable training"
            ]):
                bucket_hits["infra_prod"] += 1

            if any(k in tc for k in [
                "optimization", "optimization theory", "regularization",
                "linear algebra", "convex optimization", "non-convex optimization",
                "stochastic optimization", "gradient methods", "numerical optimization"
            ]):
                bucket_hits["math_optimization"] += 1

            # Expanded framework/system tooling bucket
            if any(k in tc for k in [
                "pytorch", "jax", "tensorflow",
                "automatic differentiation", "autodiff", "autograd",
                "computational graph", "dynamic graph", "static graph",
                "jit", "xla", "graph compiler", "compiler", "runtime",
                "kernel fusion", "operator fusion", "custom operator", "operator",
                "tensor library", "tensor computation", "vectorization",
                "dataloader", "data pipeline",
                "distributed training", "parameter server", "all-reduce",
                "mixed precision", "fp16", "bf16", "quantization",
                "onnx", "tensorrt", "triton", "xgboost", "lightgbm",
                "mlir", "tvm", "torchscript", "tensorflow serving"
            ]):
                bucket_hits["ml_frameworks"] += 1

            if any(k in tc for k in [
                "reproducible research", "research codebase", "experimental framework",
                "experiment tracking", "ablation study", "benchmarking",
                "software engineering", "testing", "modular code"
            ]):
                bucket_hits["research_workflows"] += 1

            if any(k in tc for k in [
                "state of the art", "sota", "empirical study",
                "comparative analysis", "failure modes",
                "trade-offs", "design decisions", "limitations",
                "alternative approach", "rethinking"
            ]):
                bucket_hits["inventive_thinking"] += 1

            # Software / artifact bucket
            if any(k in tc for k in [
                "open source", "github", "library", "toolkit", "framework",
                "api", "sdk", "package", "software", "implementation",
                "system", "platform", "pipeline", "serving"
            ]):
                bucket_hits["software_artifacts"] += 1

        scored_works.sort(reverse=True, key=lambda x: (x[0], x[1]))
        top = scored_works[:10]
        base_score = sum(x[0] for x in top)
        coverage = sum(1 for v in bucket_hits.values() if v > 0)

        # Neighborhood bonus: authors repeatedly appearing across the union neighborhood
        neighborhood_bonus = min(author_counter[aid], 10)
        author_score = base_score + (coverage * 3) + neighborhood_bonus

        # FIX: Instead of crashing / KeyError, apply a penalty if no artifact signal
        if bucket_hits["software_artifacts"] == 0:
            skipped_by_artifacts += 1
            author_score += -5  # penalty (keeps them but pushes down)

        top_titles = " | ".join([f"{t[2]}:{t[3]}" for t in top[:3] if t[3]])
        cited_by_appearances = author_counter[aid]

        # NEW: GitHub sniff test (optional)
        if GITHUB_SNIFF_ENABLED:
            github_checked += 1
            github_found, github_industry_signal, github_notes = github_sniff_for_author(name)
            if github_found:
                github_found_count += 1
            report_progress("github_sniff", github_checked, github_total_est, f"GitHub sniff: {github_checked}/{github_total_est}")
        else:
            github_found, github_industry_signal, github_notes = False, 0, "github_sniff_disabled"

        rows.append({
            "author_name": name,
            "author_openalex": author.get("id", ""),
            "org": org,
            "author_score": author_score,
            "bucket_coverage": coverage,
            "citing_papers_count": cited_by_appearances,

            "infra_prod_hits": bucket_hits["infra_prod"],
            "math_optimization_hits": bucket_hits["math_optimization"],
            "ml_frameworks_hits": bucket_hits["ml_frameworks"],
            "research_workflows_hits": bucket_hits["research_workflows"],
            "inventive_thinking_hits": bucket_hits["inventive_thinking"],
            "software_artifacts_hits": bucket_hits["software_artifacts"],

            # NEW: GitHub columns
            "github_found": github_found,
            "github_industry_signal": github_industry_signal,
            "github_notes": github_notes,

            "example_citing_titles": " | ".join(author_to_example_works[aid]),
            "top_titles": top_titles,
            "linkedin_xray": make_linkedin_xray(name, org),
            "eng_gate_score": eng_score,
        })

        kept += 1
        if idx % 25 == 0:
            print(
                f"Processed {idx}/{len(author_pool)} authors... "
                f"kept={kept}, skipped_by_gate={skipped_by_gate}, skipped_by_artifacts={skipped_by_artifacts}, "
                f"github_checked={github_checked}, github_found={github_found_count}"
            )
            report_progress("authors", idx, len(author_pool), f"Processed authors: {idx}/{len(author_pool)}")

    print(
        f"Done. kept={kept}, skipped_by_gate={skipped_by_gate}, skipped_by_artifacts={skipped_by_artifacts}, "
        f"github_checked={github_checked}, github_found={github_found_count}"
    )

    # 4) Export XLSX (same columns/rows as CSV, with author_name as hyperlink to Google)

    # Sort: industry signal first, then author_score (descending)
    rows.sort(
        key=lambda r: (
            r.get("github_industry_signal", 0),
            r.get("author_score", 0)
        ),
        reverse=True
    )

    out = "openalex_ml_perf_candidates_multiseed.xlsx"

    if not rows:
        fieldnames = [
            "author_name", "author_openalex", "org", "author_score", "bucket_coverage", "citing_papers_count",
            "infra_prod_hits", "math_optimization_hits", "ml_frameworks_hits", "research_workflows_hits",
            "inventive_thinking_hits", "software_artifacts_hits",
            "github_found", "github_industry_signal", "github_notes",
            "example_citing_titles", "top_titles", "linkedin_xray", "eng_gate_score"
        ]
        wb = Workbook()
        ws = wb.active
        ws.title = "candidates"

        # Capitalized + bold headers
        header_font = Font(bold=True)
        header_display = [h.replace("_", " ").upper() for h in fieldnames]
        ws.append(header_display)
        for cell in ws[1]:
            cell.font = header_font

        report_progress("export", 1, 1, "Exporting XLSX…")
        wb.save(out)
        print(f"Wrote (empty) {out} — all candidates filtered out. Loosen gates or add more/better seeds.")
        return

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "candidates"

    headers = list(rows[0].keys())

    # Capitalized + bold headers (display only; data columns stay in same order)
    header_font = Font(bold=True)
    header_display = [h.replace("_", " ").upper() for h in headers]
    ws.append(header_display)
    for cell in ws[1]:
        cell.font = header_font

    # Write rows
    for r in rows:
        ws.append([r.get(h, "") for h in headers])

    # Turn author_name cells into hyperlinks (Google search for that name)
    name_col = headers.index("author_name") + 1
    for row_idx in range(2, ws.max_row + 1):
        name_cell = ws.cell(row=row_idx, column=name_col)
        nm = name_cell.value or ""
        if nm:
            url = make_google_name_search(str(nm))
            name_cell.hyperlink = url
            name_cell.style = "Hyperlink"

    # Optional: light column width sizing (keeps it readable)
    for col_idx, h in enumerate(headers, 1):
        max_len = len(str(h))
        for row_idx in range(2, min(ws.max_row, 200) + 1):
            v = ws.cell(row=row_idx, column=col_idx).value
            if v is None:
                continue
            max_len = max(max_len, min(len(str(v)), 80))
        ws.column_dimensions[get_column_letter(col_idx)].width = max(12, min(max_len + 2, 60))

    report_progress("export", 1, 1, "Exporting XLSX…")
    wb.save(out)
    print(f"Wrote {out}")


if __name__ == "__main__":
    main()

# IMPORTANT:
# - Paste your FULL working OpenAlex/GitHub pipeline script BELOW.
# - Keep ALL comments, POS/NEG terms, etc. EXACTLY the same.
# - Only requirements:
#   (1) It must define a function: main()
#   (2) It must export XLSX to: openalex_ml_perf_candidates_multiseed.xlsx
#
# ----------------- START PASTE BELOW THIS LINE -----------------
